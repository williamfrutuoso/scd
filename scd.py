import pandas as pd
import openpyxl
from datetime import datetime

def scd():

    # Criar DataFrame 1 para buscar os dados do banco OTLP
    df_oltp_usuario = pd.read_excel('data\oltp_usuario.xlsx', sheet_name='oltp')

     # Criando DataFrame 2 para buscar os dados do banco DW
    df_dw_dusuario = pd.read_excel('data\dUsuario.xlsx', sheet_name='Planilha1', dtype={'DATA_TO': 'datetime64[ns]', 'DATA_FROM': 'datetime64[ns]'})

    #Filtrando Apenas os dados Ativo do Dataframe 2, pois os dados inativo são as versões antigas onde não precisamos fazer verificação
    df_dw_dusuario = df_dw_dusuario.loc[df_dw_dusuario['ATIVO'] == 1]

    # Criando Lista com os CPF existentes no DW dUsuario, esse campo sera utilizado como chave unica da tabela para fazer verificação
    lista_cpf = df_dw_dusuario['CPF'].to_list()

    #Verificar se o CPF que esta na base OTPL esta no DW
    for cpf in df_oltp_usuario['CPF']:
        if cpf in lista_cpf:
            #Caso caia aqui, precisamos verificar se houve alteração no dado

            #Criaremos dois DataFrame para trazer os dados do CPF que caiu nessa condição, porque dois ? pois 1 sera com os dados do banco otlp e o outro com o dados do dw
            row_oltp = df_oltp_usuario[df_oltp_usuario['CPF'] == cpf]
            row_dw = df_dw_dusuario[df_dw_dusuario['CPF'] == cpf]

            #Carregar campos para verificação
            config_campos = pd.read_excel('data\config.xlsx',sheet_name='Planilha1')
            
            #Iremos filtrar apenas as colunas que ser houver alteração precisaremos fazer um INSERT após isso criar uma lista com os campos
            campos_insert = config_campos.loc[config_campos['Tipo'] == 'insert']
            lista_insert = campos_insert['Coluna'].to_list()

            #Iremos filtrar apenas as colunas que ser houver alteração precisaremos fazer um UPDATE após isso criar uma lista com os campos
            campos_update= config_campos.loc[config_campos['Tipo'] == 'update']
            lista_update = campos_update['Coluna'].to_list()
            
            #Criaremos essa variavel para verificar se iremos a realizada a verificação de update de dados
            update = True

            #Iremos percorrer por cada coluna para saber se houve alteração
            for campo in lista_insert:
                if (row_oltp[campo].values[0] != row_dw[campo].values[0]):

                    #Como iremos inserir uma nova linha precisamos alterar os campos DATE_FROM e ATIVO antes de inserir a nova linha

                    #Iremos localizar no DataFrame 2 DW qual a possição do item que iremos alterar
                    idx = df_dw_dusuario.loc[df_dw_dusuario['CPF'] == cpf].index[0]

                    #Iremos carregar o arquivo para podermos altera-lo
                    workbook = openpyxl.load_workbook('data\dUsuario.xlsx')

                    # Selecionar o nome da aba que iremos fazer alteração
                    sheet = workbook['Planilha1']

                    #Iremos criar uma variavel data para podermos trabalhar com o DATA_FROM e DATA_TO do DataFrame 2
                    data_atual = datetime.now()

                    #Convertermos a data para um novo formato
                    data_atual = data_atual.strftime("%Y-%m-%d %H:%M:%S")
                    data_atual = str(data_atual)

                    # Iremos Inativar o registro colocando 0 na coluna ATIVO e colocando a data de finalização do dado no DATA_TO
                    sheet.cell(row=idx + 2, column=df_dw_dusuario.columns.get_loc('ATIVO') + 1, value=0)
                    sheet.cell(row=idx + 2, column=df_dw_dusuario.columns.get_loc('DATA_TO') + 1, value=data_atual)

                    #Feito isso iremos monta um DataFrame para inserir os dados no na base
                    #Iremos fazer uma nova leitura no DataFrame 2 pra buscar o ultimo ID inserido
                    buscar_ultimos_ids = pd.read_excel('data\dUsuario.xlsx', sheet_name='Planilha1', dtype={'DATA_TO': 'datetime64[ns]', 'DATA_FROM': 'datetime64[ns]'})
                    
                    # Buscamos o Ultimo ID da coluna ID do DataFrame
                    id = buscar_ultimos_ids['ID'].iloc[-1] + 1

                    # Buscamos a ultima versão do dado do participante e somamos + 1
                    version = row_dw['VERSION'].values[0] + 1
                    
                    #Preenchemos abaixo as variaveis para montar o Dataframe
                    data_from = data_atual
                    data_to = datetime.strptime('2200-01-01 00:00:00', "%Y-%m-%d %H:%M:%S")
                    data_to = str(data_to)
                    nome = row_oltp['NOME']
                    cpf = row_oltp['CPF']
                    telefone = row_oltp['TELEFONE']
                    cep = row_oltp['CEP']
                    endereco = row_oltp['ENDERECO']
                    numero = row_oltp['NUMERO']
                    ativo = 1

                    insert = pd.DataFrame({'ID': id, 'VERSION': version, 'DATA_FROM': data_from, 'DATA_TO': data_to, 'CPF': cpf,
                                      'NOME': nome, 'TELEFONE': telefone, 'CEP': cep, 'ENDERECO': endereco, 'NUMERO': numero, 'ATIVO': ativo})

                    # Converta o DataFrame em uma lista de listas
                    data_list = insert.values.tolist()

                    # Insira os dados na planilha
                    for row in data_list:
                        sheet.append(row)

                    # Salve as alterações
                    workbook.save('data\dUsuario.xlsx')

                    print(f"O registro do/da {row_oltp['NOME'].values[0]} foi versionado devido alteração no campo principal")

                                       

                    #Setamos variavel para false, pois se houve alteração no dado e precisa fazer o insert, não é necessario veriricar a proxima etapa
                    update = False
                    break
            
            if update == True:
                for campo in lista_update:
                    if (row_oltp[campo].values[0] != row_dw[campo].values[0]):
                        
                        #Iremos localizar no DataFrame 2 DW qual a possição do item que iremos alterar
                        idx = df_dw_dusuario.loc[df_dw_dusuario['CPF'] == cpf].index[0]

                        #Iremos carregar o arquivo para podermos altera-lo
                        workbook = openpyxl.load_workbook('data\dUsuario.xlsx')

                        # Selecionar o nome da aba que iremos fazer alteração
                        sheet = workbook['Planilha1']

                        #Iremos criar uma variavel data para podermos trabalhar com o DATA_FROM e DATA_TO do DataFrame 2
                        data_atual = datetime.now()

                        #Convertermos a data para um novo formato
                        data_atual = data_atual.strftime("%Y-%m-%d %H:%M:%S")

                        #Iremos agora fazer o update do campo que houve alteração
                        sheet.cell(row=idx + 2, column=df_dw_dusuario.columns.get_loc(campo) + 1, value = row_oltp[campo].values[0])

                        # Salve as alterações
                        workbook.save('data\dUsuario.xlsx')

                        print(f"Dados do/da {row_oltp['NOME'].values[0]} atualizado com sucesso")

            
        else:
            #Se não existe precisamos cadastra-lo na base

            #Criaremos dois DataFrame para trazer os dados do CPF que caiu nessa condição
            row_oltp = df_oltp_usuario[df_oltp_usuario['CPF'] == cpf]

            #Iremos fazer uma nova leitura no DataFrame 2 pra buscar o ultimo ID inserido      
            df = pd.read_excel('data\dUsuario.xlsx', sheet_name='Planilha1', dtype={'DATA_TO': 'datetime64[ns]', 'DATA_FROM': 'datetime64[ns]'})

            #Iremos fazer uma verificação para caso não tenha registro na tabela, partipos do ID 1            
            if not df.empty:
                # Obtenha o último valor da coluna 'ID'
                id = df['ID'].iloc[-1] + 1
            else:
                id = 1

            #Preenchemos abaixo as variaveis para montar o Dataframe        
            version = 1
            data_from = datetime.strptime('1900-01-01 00:00:00', "%Y-%m-%d %H:%M:%S")
            data_from = str(data_from)
            data_to =  datetime.strptime('2200-01-01 00:00:00', "%Y-%m-%d %H:%M:%S")
            data_to = str(data_to)
            nome = row_oltp['NOME']
            cpf = row_oltp['CPF']
            telefone = row_oltp['TELEFONE']
            cep = row_oltp['CEP']
            endereco = row_oltp['ENDERECO']
            numero = row_oltp['NUMERO']
            ativo = 1

            insert = pd.DataFrame({'ID': id, 'VERSION': version, 'DATA_FROM': data_from, 'DATA_TO': data_to, 'CPF': cpf,
                                  'NOME': nome, 'TELEFONE': telefone, 'CEP': cep, 'ENDERECO': endereco, 'NUMERO': numero, 'ATIVO': ativo})

            # Abra o arquivo existente
            workbook = openpyxl.load_workbook('data\dUsuario.xlsx')

            # Selecione a planilha em que você deseja inserir o DataFrame
            sheet = workbook['Planilha1']

            # Converta o DataFrame em uma lista de listas
            data_list = insert.values.tolist()

            # Insira os dados na planilha
            print(f"Usuário/a {row_oltp['NOME'].values[0]} cadastrodo/a com sucesso")
            for row in data_list:
                sheet.append(row)

            # Salve as alterações
            workbook.save('data\dUsuario.xlsx')

    print('SCD da Tabela Usuario Finalizado')


if __name__ == '__main__':
    scd()